"""
Excel to Database Migration Script
===================================

Import data from Tovito Excel tracker (v4.3 or v5.0) into the database.

Usage:
    python scripts/migrate_from_excel.py --file "path/to/excel.xlsm"
"""

import warnings
warnings.filterwarnings('ignore', category=DeprecationWarning)

import sys
import os
from datetime import datetime, date
from pathlib import Path

# Add project root to Python path
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.database.models import Database, Investor, DailyNAV, Transaction, TaxEvent
from src.utils.safe_logging import get_safe_logger

logger = get_safe_logger(__name__)

try:
    import openpyxl
except ImportError:
    print("❌ Missing required package: openpyxl")
    print("   Install with: pip install openpyxl")
    sys.exit(1)


class ExcelMigrator:
    """Migrate data from Excel to database"""
    
    def __init__(self, excel_file: str):
        """
        Initialize migrator
        
        Args:
            excel_file: Path to Excel file
        """
        self.excel_file = excel_file
        self.db = Database()
        self.workbook = None
        self.stats = {
            'investors': 0,
            'daily_nav': 0,
            'transactions': 0,
            'tax_events': 0
        }
    
    def load_workbook(self):
        """Load Excel workbook"""
        try:
            print(f"📂 Loading Excel file: {self.excel_file}")
            self.workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            print(f"✅ Workbook loaded successfully")
            print(f"   Sheets: {', '.join(self.workbook.sheetnames)}")
            return True
        except FileNotFoundError:
            print(f"❌ File not found: {self.excel_file}")
            return False
        except Exception as e:
            print(f"❌ Error loading workbook: {str(e)}")
            return False
    
    def import_investors(self):
        """Import investors from Investors sheet"""
        print("\n📊 Importing Investors...")
        
        try:
            sheet = self.workbook['Investors']
        except KeyError:
            print("⚠️  'Investors' sheet not found")
            return False
        
        session = self.db.get_session()
        
        try:
            # Determine if it's a table or regular sheet
            # v5.0 uses Excel Table starting at row 5
            # v4.3 uses regular range starting at row 2
            
            # Try v5.0 first (table with headers in row 4)
            header_row = 4
            if sheet.cell(4, 1).value == 'Investor ID':
                start_row = 5
                print("   Detected: v5.0 format (Excel Table)")
            # Try v4.3 (headers in row 1)
            elif sheet.cell(1, 1).value == 'Investor ID':
                start_row = 2
                header_row = 1
                print("   Detected: v4.3 format")
            else:
                print("❌ Cannot detect Excel format")
                return False
            
            # Read investors until we hit a blank row or totals
            row = start_row
            while True:
                investor_id = sheet.cell(row, 1).value  # Column A
                
                # Stop if blank or "TOTAL"
                if not investor_id or str(investor_id).upper() == 'TOTAL':
                    break
                
                name = sheet.cell(row, 2).value  # Column B
                initial_capital = sheet.cell(row, 3).value  # Column C
                join_date_val = sheet.cell(row, 4).value  # Column D
                status = sheet.cell(row, 5).value  # Column E
                current_shares = sheet.cell(row, 6).value  # Column F
                net_investment = sheet.cell(row, 7).value  # Column G
                
                # Parse join date
                if isinstance(join_date_val, datetime):
                    join_date = join_date_val.date()
                elif isinstance(join_date_val, date):
                    join_date = join_date_val
                elif isinstance(join_date_val, str) and len(join_date_val) == 8:
                    # Format: yyyymmdd
                    join_date = date(
                        int(join_date_val[:4]),
                        int(join_date_val[4:6]),
                        int(join_date_val[6:8])
                    )
                else:
                    # Try to extract from investor_id
                    if '-' in str(investor_id):
                        date_part = str(investor_id).split('-')[0]
                        join_date = date(
                            int(date_part[:4]),
                            int(date_part[4:6]),
                            int(date_part[6:8])
                        )
                    else:
                        join_date = date.today()
                
                # Create investor
                investor = Investor(
                    investor_id=str(investor_id),
                    name=str(name),
                    initial_capital=float(initial_capital or 0),
                    join_date=join_date,
                    status=str(status or 'Active'),
                    current_shares=float(current_shares or 0),
                    net_investment=float(net_investment or 0)
                )
                
                session.add(investor)
                self.stats['investors'] += 1
                
                print(f"   ✓ Imported: Investor *** (ID: {investor_id})")
                
                row += 1
            
            session.commit()
            print(f"✅ Imported {self.stats['investors']} investors")
            return True
            
        except Exception as e:
            session.rollback()
            print(f"❌ Error importing investors: {str(e)}")
            logger.error("Investor import failed", error=str(e))
            return False
        finally:
            session.close()
    
    def import_daily_nav(self):
        """Import daily NAV history"""
        print("\n📈 Importing Daily NAV...")
        
        try:
            sheet = self.workbook['Daily NAV']
        except KeyError:
            print("⚠️  'Daily NAV' sheet not found")
            return False
        
        session = self.db.get_session()
        
        try:
            # Headers in row 1, data starts row 2
            row = 2
            imported = 0
            
            while True:
                date_val = sheet.cell(row, 1).value  # Column A
                
                # Stop at blank
                if not date_val:
                    break
                
                # Parse date
                if isinstance(date_val, datetime):
                    nav_date = date_val.date()
                elif isinstance(date_val, date):
                    nav_date = date_val
                else:
                    row += 1
                    continue
                
                nav_per_share = sheet.cell(row, 2).value  # Column B
                total_value = sheet.cell(row, 3).value  # Column C
                total_shares = sheet.cell(row, 4).value  # Column D
                daily_change = sheet.cell(row, 5).value  # Column E
                daily_change_pct = sheet.cell(row, 6).value  # Column F
                
                # Skip if no data
                if not total_value:
                    row += 1
                    continue
                
                # Create NAV record
                nav = DailyNAV(
                    date=nav_date,
                    nav_per_share=float(nav_per_share or 0),
                    total_portfolio_value=float(total_value or 0),
                    total_shares=float(total_shares or 0),
                    daily_change_dollars=float(daily_change or 0),
                    daily_change_percent=float(daily_change_pct or 0),
                    source='Imported'
                )
                
                session.add(nav)
                imported += 1
                
                row += 1
            
            session.commit()
            self.stats['daily_nav'] = imported
            print(f"✅ Imported {imported} NAV records")
            return True
            
        except Exception as e:
            session.rollback()
            print(f"❌ Error importing NAV: {str(e)}")
            logger.error("NAV import failed", error=str(e))
            return False
        finally:
            session.close()
    
    def import_transactions(self):
        """Import transaction history"""
        print("\n💰 Importing Transactions...")
        
        try:
            sheet = self.workbook['Transactions']
        except KeyError:
            print("⚠️  'Transactions' sheet not found")
            return False
        
        session = self.db.get_session()
        
        try:
            # Headers in row 1, data starts row 2
            row = 2
            imported = 0
            
            while True:
                date_val = sheet.cell(row, 1).value  # Column A
                
                # Stop at blank
                if not date_val:
                    break
                
                # Parse date
                if isinstance(date_val, datetime):
                    trans_date = date_val.date()
                elif isinstance(date_val, date):
                    trans_date = date_val
                else:
                    row += 1
                    continue
                
                investor_id = sheet.cell(row, 2).value  # Column B
                # investor_name = sheet.cell(row, 3).value  # Column C (not used)
                trans_type = sheet.cell(row, 4).value  # Column D
                amount = sheet.cell(row, 5).value  # Column E
                share_price = sheet.cell(row, 6).value  # Column F
                shares = sheet.cell(row, 7).value  # Column G
                notes = sheet.cell(row, 8).value  # Column H
                
                # Skip if no investor_id
                if not investor_id:
                    row += 1
                    continue
                
                # Create transaction
                transaction = Transaction(
                    date=trans_date,
                    investor_id=str(investor_id),
                    transaction_type=str(trans_type),
                    amount=float(amount or 0),
                    share_price=float(share_price or 0),
                    shares_transacted=float(shares or 0),
                    notes=str(notes) if notes else None
                )
                
                session.add(transaction)
                imported += 1
                
                row += 1
            
            session.commit()
            self.stats['transactions'] = imported
            print(f"✅ Imported {imported} transactions")
            return True
            
        except Exception as e:
            session.rollback()
            print(f"❌ Error importing transactions: {str(e)}")
            logger.error("Transaction import failed", error=str(e))
            return False
        finally:
            session.close()
    
    def import_tax_events(self):
        """Import tax events"""
        print("\n📋 Importing Tax Events...")
        
        try:
            sheet = self.workbook['Tax Tracker']
        except KeyError:
            print("⚠️  'Tax Tracker' sheet not found - skipping")
            return True  # Not critical
        
        session = self.db.get_session()
        
        try:
            # Find where tax events start (after the tax rate section)
            # Usually around row 5-10
            start_row = None
            for r in range(1, 20):
                cell_val = sheet.cell(r, 1).value
                if cell_val and 'Date' in str(cell_val):
                    start_row = r + 1
                    break
            
            if not start_row:
                print("⚠️  Tax events section not found - skipping")
                return True
            
            row = start_row
            imported = 0
            
            while True:
                date_val = sheet.cell(row, 1).value
                
                if not date_val:
                    break
                
                # Parse date
                if isinstance(date_val, datetime):
                    tax_date = date_val.date()
                elif isinstance(date_val, date):
                    tax_date = date_val
                else:
                    row += 1
                    continue
                
                investor_id = sheet.cell(row, 2).value
                withdrawal_amt = sheet.cell(row, 4).value
                realized_gain = sheet.cell(row, 5).value
                tax_due = sheet.cell(row, 6).value
                net_proceeds = sheet.cell(row, 7).value
                
                if not investor_id or not withdrawal_amt:
                    row += 1
                    continue
                
                # Create tax event
                tax_event = TaxEvent(
                    date=tax_date,
                    investor_id=str(investor_id).split('-')[0] if '-' in str(investor_id) else str(investor_id),
                    withdrawal_amount=float(withdrawal_amt or 0),
                    realized_gain=float(realized_gain or 0),
                    tax_due=float(tax_due or 0),
                    net_proceeds=float(net_proceeds or 0),
                    tax_rate=0.37
                )
                
                session.add(tax_event)
                imported += 1
                
                row += 1
            
            session.commit()
            self.stats['tax_events'] = imported
            print(f"✅ Imported {imported} tax events")
            return True
            
        except Exception as e:
            session.rollback()
            print(f"⚠️  Error importing tax events: {str(e)}")
            # Not critical, continue
            return True
        finally:
            session.close()
    
    def verify_import(self):
        """Verify imported data"""
        print("\n🔍 Verifying Import...")
        
        session = self.db.get_session()
        
        try:
            # Check investors
            investors = session.query(Investor).all()
            print(f"   Investors in database: {len(investors)}")
            
            # Check NAV records
            nav_records = session.query(DailyNAV).all()
            print(f"   NAV records: {len(nav_records)}")
            
            # Check transactions
            transactions = session.query(Transaction).all()
            print(f"   Transactions: {len(transactions)}")
            
            # Verify latest NAV
            latest_nav = session.query(DailyNAV).order_by(DailyNAV.date.desc()).first()
            if latest_nav:
                print(f"\n   Latest NAV:")
                print(f"   Date: {latest_nav.date}")
                print(f"   NAV per Share: $***")
                print(f"   Portfolio Value: $***")
                print(f"   Total Shares: ***")
            
            # Verify investor totals
            total_shares = sum(inv.current_shares for inv in investors if inv.status == 'Active')
            print(f"\n   Total Active Shares: ***")
            
            if latest_nav and abs(total_shares - latest_nav.total_shares) > 0.01:
                print(f"   ⚠️  Warning: Share mismatch!")
                print(f"   Investors total: ***")
                print(f"   NAV record: ***")
                return False
            
            print(f"\n✅ Verification passed!")
            return True
            
        finally:
            session.close()
    
    def run(self):
        """Run complete migration"""
        print("\n" + "="*60)
        print("EXCEL TO DATABASE MIGRATION")
        print("="*60)
        
        # Load workbook
        if not self.load_workbook():
            return False
        
        # Import each section
        success = True
        success = self.import_investors() and success
        success = self.import_daily_nav() and success
        success = self.import_transactions() and success
        success = self.import_tax_events() and success
        
        # Verify
        if success:
            success = self.verify_import()
        
        # Summary
        print("\n" + "="*60)
        if success:
            print("✅ MIGRATION COMPLETE!")
        else:
            print("⚠️  MIGRATION COMPLETED WITH WARNINGS")
        print("="*60)
        
        print(f"\nImported:")
        print(f"  • Investors: {self.stats['investors']}")
        print(f"  • Daily NAV Records: {self.stats['daily_nav']}")
        print(f"  • Transactions: {self.stats['transactions']}")
        print(f"  • Tax Events: {self.stats['tax_events']}")
        print()
        
        return success


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Migrate Excel data to database')
    parser.add_argument('--file', '-f', required=True, help='Path to Excel file')
    parser.add_argument('--clear', action='store_true', help='Clear existing data first (DANGEROUS!)')
    
    args = parser.parse_args()
    
    # Warn if clearing
    if args.clear:
        response = input("⚠️  This will DELETE all existing data! Are you sure? (yes/NO): ")
        if response.lower() != 'yes':
            print("Cancelled")
            sys.exit(0)
        
        db = Database()
        print("Clearing existing data...")
        db.drop_all_tables()
        db.create_all_tables()
        print("✅ Data cleared")
    
    # Run migration
    migrator = ExcelMigrator(args.file)
    success = migrator.run()
    
    sys.exit(0 if success else 1)
